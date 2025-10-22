from flask import session
import threading
import time
import json
import tempfile

# In-memory progress store (for demo; use Redis or DB for production)
progress_store = {}
progress_store_lock = threading.Lock()

import os
from flask import Flask, request, send_from_directory, jsonify, render_template_string
from werkzeug.utils import secure_filename
import csv
import re
from collections import defaultdict
import io as _io
# Heavy libraries (pandas, PyMuPDF, pytesseract, matplotlib, openpyxl, Pillow)
# are imported lazily inside the functions that need them to reduce startup
# memory and CPU usage on constrained hosts.
import uuid
import shutil
import traceback
import subprocess
import logging
import platform
import sys
import psutil
import boto3
from botocore.exceptions import ClientError
try:
    from redis import Redis
    from rq import Queue
except Exception:
    Redis = None
    Queue = None

# Use absolute paths for directories to avoid issues in cloud environments
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
PROGRESS_DIR = os.path.join(BASE_DIR, 'progress')
os.makedirs(PROGRESS_DIR, exist_ok=True)
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
EXPORTS_FOLDER = os.path.join(BASE_DIR, 'exports')
ALLOWED_EXTENSIONS = {'pdf', 'csv', 'xlsx'}

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(EXPORTS_FOLDER, exist_ok=True)

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['EXPORTS_FOLDER'] = EXPORTS_FOLDER
# Default to a smaller upload limit for low-performance free hosts (5MB)
app.config['MAX_CONTENT_LENGTH'] = int(os.environ.get('MAX_CONTENT_LENGTH', 5 * 1024 * 1024))  # 5MB default

# Logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger('qa-validator')

# Try to configure Redis if provided (Render/Production)
REDIS_URL = os.environ.get('REDIS_URL')
redis_conn = None
rq_queue = None
if REDIS_URL:
    try:
        redis_conn = Redis.from_url(REDIS_URL)
        rq_queue = Queue('default', connection=redis_conn)
        logger.info('Connected to Redis for background jobs')
    except Exception as e:
        logger.exception('Failed to connect to Redis: %s', e)


def set_progress(progress_key, percent=None, csv_filename=None, done=None, error=None):
    """Set progress in Redis if configured, otherwise write atomic JSON file (visible to all workers).

    Parameters:
    - progress_key: str
    - percent: int
    - csv_filename: str
    - done: bool
    - error: dict or None (e.g. {'code':'PARSE_ERROR','message':'short message'})
    """
    if redis_conn:
        data = {}
        if percent is not None:
            data['percent'] = int(percent)
        if csv_filename is not None:
            data['csv_filename'] = csv_filename
        if done is not None:
            data['done'] = int(bool(done))
        if error is not None:
            # store structured error as JSON string
            try:
                data['error'] = json.dumps(error)
            except Exception:
                data['error'] = json.dumps({'message': str(error)})
        if data:
            try:
                # Use a pipeline to ensure atomic execution when available
                pipe = redis_conn.pipeline()
                pipe.hset(progress_key, mapping={k: str(v) for k, v in data.items()})
                pipe.execute()
            except Exception:
                # Fallback to single hset
                try:
                    redis_conn.hset(progress_key, mapping={k: str(v) for k, v in data.items()})
                except Exception:
                    logger.exception('Failed to write progress to Redis for key %s', progress_key)
    else:
        # update in-memory for quick access and also persist to disk for cross-process visibility
        with progress_store_lock:
            prog = progress_store.get(progress_key)
            if not prog:
                prog = {'percent': 0, 'done': False, 'csv_filename': None}
                progress_store[progress_key] = prog
            if percent is not None:
                prog['percent'] = int(percent)
            if csv_filename is not None:
                prog['csv_filename'] = csv_filename
            if done is not None:
                prog['done'] = bool(done)
            if error is not None:
                prog['error'] = error
            # persist atomically
            path = os.path.join(PROGRESS_DIR, f"{progress_key}.json")
            tmpfd, tmppath = tempfile.mkstemp(dir=PROGRESS_DIR)
            try:
                with os.fdopen(tmpfd, 'w') as t:
                    json.dump(prog, t)
                    t.flush()
                    os.fsync(t.fileno())
                os.replace(tmppath, path)
            except Exception:
                logger.exception('Failed to write progress file for key %s', progress_key)
                try:
                    if os.path.exists(tmppath):
                        os.remove(tmppath)
                except Exception:
                    pass

def get_progress_data(progress_key):
    """Retrieve progress dict from Redis or from on-disk JSON fallback, otherwise in-memory."""
    if redis_conn:
        try:
            h = redis_conn.hgetall(progress_key)
            if not h:
                return {'percent': 0, 'done': False, 'csv_filename': None}
            decoded = {k.decode() if isinstance(k, bytes) else k: v.decode() if isinstance(v, bytes) else v for k, v in h.items()}
            err = decoded.get('error')
            parsed_err = None
            if err:
                try:
                    parsed_err = json.loads(err)
                except Exception:
                    parsed_err = {'message': err}
            return {
                'percent': int(decoded.get('percent', 0)),
                'done': bool(int(decoded.get('done', 0))) if decoded.get('done') is not None else False,
                'csv_filename': decoded.get('csv_filename'),
                'error': parsed_err
            }
        except Exception:
            logger.exception('Error reading progress from Redis for key %s', progress_key)
            return {'percent': 0, 'done': False, 'csv_filename': None}
    else:
        # Prefer the on-disk JSON (cross-process); fall back to in-memory
        path = os.path.join(PROGRESS_DIR, f"{progress_key}.json")
        try:
            if os.path.exists(path):
                with open(path, 'r') as f:
                    data = json.load(f)
                    return {
                        'percent': int(data.get('percent', 0)),
                        'done': bool(data.get('done', False)),
                        'csv_filename': data.get('csv_filename'),
                        'error': data.get('error')
                    }
        except Exception:
            logger.exception('Failed to read progress file for key %s', progress_key)
        with progress_store_lock:
            p = progress_store.get(progress_key, {'percent': 0, 'done': False, 'csv_filename': None, 'error': None})
            # ensure keys exist
            if 'error' not in p:
                p['error'] = None
            return p


def upload_to_s3(local_path, bucket, key_prefix=''):
    """Upload file to S3 and return the object key."""
    s3 = boto3.client('s3')
    key = os.path.join(key_prefix, os.path.basename(local_path)) if key_prefix else os.path.basename(local_path)
    try:
        s3.upload_file(local_path, bucket, key)
        return key
    except ClientError as e:
        logger.exception('S3 upload failed: %s', e)
        return None


def presigned_url(bucket, key, expires=3600):
    s3 = boto3.client('s3')
    try:
        url = s3.generate_presigned_url('get_object', Params={'Bucket': bucket, 'Key': key}, ExpiresIn=expires)
        return url
    except ClientError as e:
        logger.exception('Presigned URL generation failed: %s', e)
        return None

# Detect tesseract binary location; actual pytesseract import/use is lazy inside OCR code
tess_path = shutil.which('tesseract')
if tess_path:
    try:
        ver = subprocess.check_output([tess_path, '--version'], stderr=subprocess.STDOUT, text=True).splitlines()[0]
        logger.info("Tesseract found: %s (%s)", tess_path, ver)
    except Exception as e:
        logger.exception("Tesseract found at %s but failed to run --version: %s", tess_path, e)
else:
    logger.warning("Tesseract binary not found on PATH; OCR fallback may fail on this host.")

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def sanitize_csv_filename(filename):
    """Sanitize a filename to be safe for local storage/URLs and ensure it ends with .csv."""
    if not filename:
        return None
    safe = secure_filename(filename)
    base, _ = os.path.splitext(safe)
    return base + '.csv'

def extract_text_with_ocr(page):
    # `page` is a PyMuPDF page object. Try direct extraction first and fall back to OCR.
    text = page.get_text()
    if text.strip():
        return text
    # Lazy-import heavy OCR deps only when needed
    try:
        from PIL import Image
        import pytesseract
    except Exception:
        logger.warning('OCR dependencies not available; returning empty text')
        return ''
    pix = page.get_pixmap(dpi=150)
    img = Image.open(_io.BytesIO(pix.tobytes("png")))
    try:
        text = pytesseract.image_to_string(img)
    finally:
        img.close()
    return text

def validate_pdf(pdf_path, export_dir, progress_key=None, result_key=None):
    base_name = os.path.splitext(os.path.basename(pdf_path))[0]
    csv_path = os.path.join(export_dir, f"{base_name}_validation_summary.csv")
    excel_path = os.path.join(export_dir, f"{base_name}_validation_summary.xlsx")
    dashboard_path = os.path.join(export_dir, f"{base_name}_dashboard.png")

    REQUIRED_FIELDS = [
        "Customer Name", "Customer P.O. Number", "Customer Part Number",
        "Customer Part Number Revision", "AEM Part Number", "AEM Lot Number",
        "AEM Date Code", "AEM Cage Code", "Customer Quality Clauses",
        "FAI Form 3", "Solderability Test Report", "DPA", "Visual Inspection Record",
        "Shipment Quantity", "Reel Labels", "Certificate of Conformance", "Route Sheet",
        "Part Number", "Lot Number", "Date", "Resistance", "Dimension", "Test Result"
    ]

    # Precompile regex patterns for all fields for speed
    FIELD_PATTERNS = {field: re.compile(rf"{re.escape(field)}[:\s]*([^\n]+)", re.IGNORECASE) for field in REQUIRED_FIELDS}

    NUMERICAL_RANGES = {
        "Resistance": (95, 105),
        "Dimension": (0.9, 1.1)
    }

    anomalies = []
    critical_issues = []
    field_presence = defaultdict(int)
    all_fields = []
    field_page_map = defaultdict(list)  # field -> list of page numbers
    field_value_map = defaultdict(list)  # field -> list of (page, value)
    # Segment settings: process PDFs in chunks to minimize memory peak for very large PDFs
    try:
        segment_size = int(os.environ.get('PAGE_SEGMENT_SIZE', '4'))
        if segment_size <= 0:
            segment_size = 4
    except Exception:
        segment_size = 4
    segment_files = []

    def extract_fields(text):
        """Extract fields by locating field label positions and taking the text up to the next label.
        This captures multi-line values and cases where a colon isn't present.
        Falls back to the simple regex if positional extraction doesn't find a value.
        """
        fields = {}
        if not text:
            return fields

        # Work with the original text for extraction but perform case-insensitive searches
        lower_text = text.lower()

        # Find all label occurrences with their span
        occurrences = []  # list of (start_index, end_index, field)
        for field in REQUIRED_FIELDS:
            low_field = field.lower()
            for m in re.finditer(re.escape(low_field), lower_text):
                occurrences.append((m.start(), m.end(), field))

        # If no positional occurrences found, fall back to simple pattern matching
        if not occurrences:
            for field, pattern in FIELD_PATTERNS.items():
                match = pattern.search(text)
                if match:
                    fields[field] = match.group(1).strip()
            return fields

        # Sort occurrences by position
        occurrences.sort(key=lambda x: x[0])

        for idx, (start, end, field) in enumerate(occurrences):
            value_start = end
            value_end = occurrences[idx + 1][0] if idx + 1 < len(occurrences) else len(text)
            raw_value = text[value_start:value_end]
            # Remove leading separators and whitespace
            raw_value = re.sub(r"^[\s:.-]*", "", raw_value)
            # Trim trailing whitespace/newlines and collapse internal whitespace
            value = re.sub(r"\s+", " ", raw_value).strip()
            # Limit to a reasonable length to avoid grabbing huge blocks
            if value:
                fields[field] = value[:1000]

        # For any required field still missing, try the regex fallback once
        for field, pattern in FIELD_PATTERNS.items():
            if field not in fields:
                match = pattern.search(text)
                if match:
                    fields[field] = match.group(1).strip()

        return fields

    def validate_numerical(field, value):
        try:
            val = float(re.findall(r"[\d.]+", value)[0])
            min_val, max_val = NUMERICAL_RANGES[field]
            return min_val <= val <= max_val
        except:
            return False

    def check_consistency(field_name):
        values = [fields.get(field_name) for fields in all_fields if field_name in fields]
        return len(set(values)) == 1

    # Lazy-import heavy libraries used for PDF processing and reporting
    try:
        import fitz  # PyMuPDF
        from PIL import Image
        import pytesseract
        import pandas as pd
        import matplotlib.pyplot as plt
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except Exception as e:
        logger.exception('Missing optional heavy dependency: %s', e)
        # re-raise so caller handles the error and produces an error CSV
        raise

    doc = fitz.open(pdf_path)
    total_pages = len(doc)
    # buffer for current segment rows (list of (page, field, status, output))
    segment_rows = []
    segment_index = 0
    for page_num in range(total_pages):
        page = doc.load_page(page_num)
        # Try to extract text directly; only use OCR if text is empty
        text = page.get_text()
        if not text.strip():
            # Only do OCR if absolutely necessary
            pix = page.get_pixmap(dpi=150)
            img = Image.open(_io.BytesIO(pix.tobytes("png")))
            try:
                text = pytesseract.image_to_string(img)
            finally:
                img.close()
        fields = extract_fields(text)
        all_fields.append(fields)

        for field in REQUIRED_FIELDS:
            if field not in fields:
                anomalies.append([page_num + 1, field, "Missing"])
            else:
                field_presence[field] += 1
                field_page_map[field].append(page_num + 1)
                field_value_map[field].append((page_num + 1, fields[field]))

        for field in NUMERICAL_RANGES:
            if field in fields and not validate_numerical(field, fields[field]):
                anomalies.append([page_num + 1, field, f"Out of range: {fields[field]}"])
                critical_issues.append([page_num + 1, field, fields[field]])

        # Prepare rows for this page into the current segment buffer
        for field in REQUIRED_FIELDS:
            if field in fields:
                segment_rows.append((page_num + 1, field, 'Found', fields[field]))
            else:
                segment_rows.append((page_num + 1, field, 'Missing', ''))

        # If segment boundary reached or last page, flush segment to disk
        if ((page_num + 1) % segment_size == 0) or (page_num == total_pages - 1):
            seg_path = os.path.join(export_dir, f"{base_name}_segment_{segment_index}_validation_summary.csv")
            try:
                with open(seg_path, 'w', newline='') as sf:
                    writer = csv.writer(sf)
                    writer.writerow(["Page", "Field", "Result", "Output"])
                    for r in segment_rows:
                        writer.writerow(r)
                segment_files.append(seg_path)
            except Exception:
                logger.exception('Failed to write segment CSV: %s', seg_path)
            # clear buffer for next segment
            segment_rows = []
            segment_index += 1

        # Update progress - but don't set 100% here to avoid race condition
        if progress_key and page_num + 1 < total_pages:
            set_progress(progress_key, percent=int(((page_num + 1) / total_pages) * 100))
        # Log progress at a coarse level
        if (page_num + 1) % 10 == 0 or page_num == total_pages - 1:
            logger.info('Processed page %s/%s for %s', page_num + 1, total_pages, base_name)

    for field in ["Part Number", "Lot Number", "Date"]:
        if not check_consistency(field):
            anomalies.append(["All Pages", field, "Inconsistent values"])
            critical_issues.append(["All Pages", field, "Inconsistent values"])

    # Merge segment CSVs into the final summary (preserve order)
    try:
        with open(csv_path, 'w', newline='') as out_f:
            writer = csv.writer(out_f)
            writer.writerow(["Page", "Field", "Result", "Output"])
            for seg in sorted(segment_files, key=lambda s: int(re.search(r"_segment_(\d+)_", os.path.basename(s)).group(1)) if re.search(r"_segment_(\d+)_", os.path.basename(s)) else 0):
                try:
                    with open(seg, 'r', newline='') as sf:
                        r = csv.reader(sf)
                        header = next(r, None)
                        for row in r:
                            writer.writerow(row)
                except Exception:
                    logger.exception('Failed to merge segment file: %s', seg)
        # Optionally remove segment files after merge
        for seg in segment_files:
            try:
                os.remove(seg)
            except Exception:
                pass
    except Exception:
        logger.exception('Failed to write final merged CSV: %s', csv_path)

    # Write summary: all required fields, 'Found' if present, and all extracted values (concatenated)
    field_info_csv = os.path.join(export_dir, f"{base_name}_field_info_summary.csv")
    with open(field_info_csv, "w", newline='') as f:
        writer = csv.writer(f)
        writer.writerow(["Field", "Status", "Output"])
        for field in REQUIRED_FIELDS:
            values = field_value_map[field]
            if values:
                # Concatenate all extracted values for this field, preserving all characters
                value_str = '; '.join(v for _, v in values)
                writer.writerow([field, "Found", value_str])
            else:
                writer.writerow([field, "Not found", "Not found"])

    wb = Workbook()
    ws = wb.active
    ws.title = "QA Anomalies"

    headers = ["Page", "Field", "Issue"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    for row_num, row_data in enumerate(anomalies, start=2):
        for col_num, cell_value in enumerate(row_data, start=1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

    table_ref = f"A1:C{len(anomalies)+1}"
    table = Table(displayName="AnomalyTable", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    wb.save(excel_path)

    plt.figure(figsize=(12, 6))
    plt.bar(field_presence.keys(), field_presence.values(), color='skyblue')
    plt.title("Field Presence Across PDF Pages")
    plt.xlabel("Field Name")
    plt.ylabel("Number of Pages Present")
    plt.xticks(rotation=90)
    plt.tight_layout()
    plt.savefig(dashboard_path)
    # If S3 configured, upload CSV and dashboard, and return presigned URL through progress store
    s3_bucket = os.environ.get('S3_BUCKET')
    s3_prefix = os.environ.get('S3_PREFIX', '')
    if s3_bucket:
        logger.info('S3 bucket configured: %s, attempting uploads', s3_bucket)
        csv_key = upload_to_s3(csv_path, s3_bucket, s3_prefix)
        dash_key = upload_to_s3(dashboard_path, s3_bucket, s3_prefix)
        field_info_key = upload_to_s3(field_info_csv, s3_bucket, s3_prefix)
        logger.info('S3 uploads complete. CSV key: %s', csv_key)
        if progress_key:
            if csv_key:
                # S3 upload succeeded, use S3 key as filename (sanitize basename)
                csvfn = sanitize_csv_filename(os.path.basename(csv_key))
                logger.info('Setting progress with S3 key: %s', csvfn)
                set_progress(progress_key, percent=100, csv_filename=csvfn, done=True)
            else:
                # S3 upload failed, fallback to local file serving
                csvfn = sanitize_csv_filename(os.path.basename(csv_path))
                logger.warning('S3 upload failed, falling back to local file serving')
                set_progress(progress_key, percent=100, csv_filename=csvfn, done=True)
        else:
            logger.warning('No progress_key provided for S3 path')
    else:
        # Save result in progress store for robust retrieval
        logger.info('No S3 bucket configured, using local file serving')
        if progress_key:
            csvfn = sanitize_csv_filename(os.path.basename(csv_path))
            logger.info('Setting progress with local filename: %s', csvfn)
            set_progress(progress_key, percent=100, csv_filename=csvfn, done=True)
        else:
            logger.warning('No progress_key provided for local path')

    logger.info('Validation complete. CSV saved at %s', csv_path)
    logger.info('Progress key: %s, S3 bucket: %s', progress_key, s3_bucket)
    return csv_path, excel_path, dashboard_path, len(anomalies), len(critical_issues)

def validate_file(filepath, progress_key=None, result_key=None):
    # If PDF, run PDF validation, else fallback to dummy
    if filepath.lower().endswith('.pdf'):
        # Run full PDF validation
        csv_path, excel_path, dashboard_path, anomaly_count, critical_count = validate_pdf(filepath, EXPORTS_FOLDER, progress_key, result_key)
        # Lazy-import pandas only when needed
        import pandas as pd
        df = pd.read_csv(csv_path)
        print(f"validate_file: CSV generated at {csv_path}")
        return df, os.path.basename(csv_path)
    else:
        # Non-PDF fallback
        import pandas as pd
        data = {'filename': [os.path.basename(filepath)], 'status': ['validated']}
        df = pd.DataFrame(data)
        csv_filename = os.path.splitext(os.path.basename(filepath))[0] + '.csv'
        csv_path = os.path.join(EXPORTS_FOLDER, csv_filename)
        df.to_csv(csv_path, index=False)
        if progress_key:
            set_progress(progress_key, percent=100, csv_filename=sanitize_csv_filename(csv_filename), done=True)
        print(f"validate_file: Non-PDF CSV generated at {csv_path}")
        return df, csv_filename

def export_to_csv(df, csv_path):
    df.to_csv(csv_path, index=False)

@app.route('/', methods=['GET'])
def index():
    return render_template_string('''
    <h2>Upload file for validation</h2>
    <p>1. Select a file.<br>
    2. Click <b>Upload and Validate</b>.<br>
    3. Wait for both progress bars to reach 100%.<br>
    4. When validation is complete, click the <b>Download CSV</b> link.</p>
        <form id="upload-form" method="post" enctype="multipart/form-data" onsubmit="return false;">
            <input type="file" name="file" id="file-input">
            <button id="upload-button" type="button">Upload and Validate</button>
            <noscript>
                <p style="color: red;">JavaScript is required for in-page progress. With no JavaScript, the form will perform a full-page POST which is not recommended. Please enable JavaScript.</p>
            </noscript>
        </form>
    <div style="margin-top:20px;">
      <div>Upload Progress: <span id="upload-percent">0%</span></div>
      <div id="upload-progress-bar" style="width: 100%; background: #eee; height: 20px;">
        <div id="upload-progress" style="background: #2196f3; width: 0%; height: 100%;"></div>
      </div>
    </div>
    <div style="margin-top:20px;">
      <div>Validation Progress: <span id="progress-percent">0%</span></div>
      <div id="progress-bar" style="width: 100%; background: #eee; height: 20px;">
        <div id="progress" style="background: #4caf50; width: 0%; height: 100%;"></div>
      </div>
    </div>
    <div id="download-link" style="margin-top:20px;"></div>
        <script>
        // Ensure the JS handler is attached; if it fails, the form will not submit and user will see an error.
        (function() {
            const form = document.getElementById('upload-form');
            const uploadButton = document.getElementById('upload-button');
            if (!form || !uploadButton) {
                console.error('Upload form or button not found; upload handler cannot be attached.');
                return;
            }

            async function startUpload() {
                const e = { preventDefault: () => {} };
                // reuse the same logic as before
                const formData = new FormData(form);
                let uploadProgressBar = document.getElementById('upload-progress');
                let uploadPercentText = document.getElementById('upload-percent');
                let progressBar = document.getElementById('progress');
                let progressPercentText = document.getElementById('progress-percent');
                uploadProgressBar.style.width = '0%';
                uploadPercentText.innerText = '0%';
                progressBar.style.width = '0%';
                progressPercentText.innerText = '0%';
                document.getElementById('download-link').innerHTML = '';

                // AJAX upload with progress
                const xhr = new XMLHttpRequest();
                xhr.open('POST', '/api/validate', true);

                xhr.upload.onprogress = function(e) {
                    if (e.lengthComputable) {
                        let percent = Math.round((e.loaded / e.total) * 100);
                        uploadProgressBar.style.width = percent + '%';
                        uploadPercentText.innerText = percent + '%';
                    }
                };

                xhr.onreadystatechange = async function() {
                    if (xhr.readyState === XMLHttpRequest.DONE) {
                        if (xhr.status === 200) {
                            uploadProgressBar.style.width = '100%';
                            uploadPercentText.innerText = '100%';
                            const data = JSON.parse(xhr.responseText);
                            if (!data.progressKey) {
                                document.getElementById('download-link').innerText = 'Validation failed.';
                                return;
                            }
                            // Poll for validation progress
                            let percent = 0;
                            let csvFilename = '';
                                            while (percent < 100) {
                                                const progRes = await fetch(`/api/progress/${data.progressKey}`);
                                                const progData = await progRes.json();
                                                percent = progData.percent;
                                                progressBar.style.width = percent + '%';
                                                progressPercentText.innerText = percent + '%';
                                                if (progData.done) {
                                                    if (progData.error) {
                                                        // Display user-friendly message if available
                                                        const code = progData.error.code || '';
                                                        const msg = progData.error.message || progData.error;
                                                        document.getElementById('download-link').innerHTML = `<div style="color: red;"><b>Error</b> ${code}: ${msg}</div>`;
                                                        csvFilename = '';
                                                        break;
                                                    }
                                                    if (progData.csv_filename) {
                                                        csvFilename = progData.csv_filename;
                                                        break;
                                                    }
                                                }
                                                await new Promise(r => setTimeout(r, 1000));
                                            }
                            progressBar.style.width = '100%';
                            progressPercentText.innerText = '100%';
                                        if (csvFilename) {
                                                document.getElementById('download-link').innerHTML =
                                                    `<b>Validation complete! Click the link below to download your results:</b><br>
                                                    <a href="/download/${csvFilename}" download>Download CSV</a>`;
                                        } else {
                                                // If no csvFilename it's either an error already rendered or a failure
                                                if (!document.getElementById('download-link').innerHTML) {
                                                        document.getElementById('download-link').innerText = 'Validation failed.';
                                                }
                                        }
                        } else {
                            document.getElementById('download-link').innerText = 'Upload failed.';
                        }
                    }
                };
                xhr.send(formData);
            }

            // Attach click handler and provide a simple error if the handler is missing later
            uploadButton.addEventListener('click', function() {
                try {
                    startUpload();
                } catch (err) {
                    console.error('Upload handler error:', err);
                    document.getElementById('download-link').innerText = 'Upload handler failed. See console.';
                }
            });
        })();
        </script>
    ''')

@app.route('/api/validate', methods=['POST'])
def api_validate():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        upload_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(upload_path)

        # Generate a unique progress key
        progress_key = str(uuid.uuid4())
        set_progress(progress_key, percent=0, csv_filename=None, done=False)

        def run_validation_local(progress_key, upload_path, filename):
            try:
                logger.info(f"Starting validation for {upload_path} with progress_key: {progress_key}")
                csv_path, excel_path, dashboard_path, anomaly_count, critical_count = validate_pdf(upload_path, EXPORTS_FOLDER, progress_key, progress_key)
                csv_filename = os.path.basename(csv_path)
                logger.info(f"Validation finished for {upload_path}, CSV: {csv_filename}")
            except Exception as e:
                logger.exception(f"Validation error for {upload_path}: {e}")
                error_csv = os.path.splitext(filename)[0] + "_validation_summary.csv"
                error_csv_path = os.path.join(EXPORTS_FOLDER, error_csv)
                try:
                    tb = traceback.format_exc()
                    with open(error_csv_path, "w", newline='') as f:
                        writer = csv.writer(f)
                        writer.writerow(["Error"])
                        writer.writerow([str(e)])
                        writer.writerow(["Traceback:"])
                        for line in tb.splitlines():
                            writer.writerow([line])
                    set_progress(progress_key, percent=100, csv_filename=sanitize_csv_filename(error_csv), done=True, error={'message': str(e)})
                    logger.info(f"Error CSV written: {error_csv_path}")
                except Exception as file_error:
                    logger.exception(f"Error writing error CSV: {file_error}")
                    set_progress(progress_key, percent=100, csv_filename=None, done=True, error={'message': str(file_error)})
            except:
                # Catch any unexpected exceptions
                logger.exception(f"Unexpected error during validation for {upload_path}")
                set_progress(progress_key, percent=100, csv_filename=None, done=True, error={'message': 'Unexpected error occurred'})

        if rq_queue:
            # Enqueue the validation job to Redis queue
            try:
                job = rq_queue.enqueue('app.validate_file', upload_path, progress_key, progress_key)
                print('Enqueued job', job.id)
            except Exception as e:
                print('Failed to enqueue job, falling back to thread:', e)
                thread = threading.Thread(target=run_validation_local, args=(progress_key, upload_path, filename))
                thread.start()
        else:
            # Run locally in background thread
            thread = threading.Thread(target=run_validation_local, args=(progress_key, upload_path, filename))
            thread.start()

        return jsonify({'progressKey': progress_key})
    return jsonify({'error': 'Invalid file type'}), 400

@app.route('/api/progress/<progress_key>', methods=['GET'])
def get_progress(progress_key):
    prog = get_progress_data(progress_key)
    logger.info('Progress check for key %s: %s', progress_key, prog)
    if not prog:
        return jsonify({'percent': 0, 'done': False, 'csv_filename': None, 'error': 'Progress key not found'}), 404
    # If percent is 100 but 'done' is not set, attempt a safe auto-fix.
    # This addresses cases where a worker/process updated percent and csv_filename
    # but failed to persist the final 'done' flag (for example due to a cross-process
    # file-write or permission issue). If the CSV file exists locally (or a filename
    # is present when S3 is configured) we mark the progress as done and persist it.
    try:
        if prog.get('percent', 0) >= 100 and not prog.get('done'):
            csvfn = prog.get('csv_filename')
            s3_bucket = os.environ.get('S3_BUCKET')
            file_exists = False
            if csvfn:
                if s3_bucket:
                    # assume S3 upload completed if a filename exists; the presigned URL
                    # generation will still validate availability.
                    file_exists = True
                else:
                    local_path = os.path.join(EXPORTS_FOLDER, csvfn)
                    if os.path.exists(local_path):
                        file_exists = True
            # If we detect the CSV is available, persist done=True so clients stop polling.
            if file_exists:
                logger.info('Auto-marking progress done for key %s since percent>=100 and file exists', progress_key)
                set_progress(progress_key, percent=prog.get('percent', 100), csv_filename=csvfn, done=True)
                # refresh prog from the authoritative source
                prog = get_progress_data(progress_key)
    except Exception:
        logger.exception('Error while attempting to auto-mark progress done for key %s', progress_key)

    # If S3 is configured and csv_filename is present, return presigned URL
    s3_bucket = os.environ.get('S3_BUCKET')
    if s3_bucket and prog.get('csv_filename'):
        s3_prefix = os.environ.get('S3_PREFIX', '')
        key = os.path.join(s3_prefix, prog.get('csv_filename')) if s3_prefix else prog.get('csv_filename')
        url = presigned_url(s3_bucket, key)
        return jsonify({
            'percent': prog.get('percent', 0),
            'done': prog.get('done', False),
            'csv_filename': prog.get('csv_filename'),
            'download_url': url,
            'error': prog.get('error')
        })

    return jsonify({
        'percent': prog.get('percent', 0),
        'done': prog.get('done', False),
        'csv_filename': prog.get('csv_filename')
        , 'error': prog.get('error')
    })

@app.route('/download/<csv_filename>', methods=['GET'])
def download_csv(csv_filename):
    # If S3 is configured, return presigned URL; otherwise serve from exports folder
    s3_bucket = os.environ.get('S3_BUCKET')
    s3_prefix = os.environ.get('S3_PREFIX', '')
    if s3_bucket:
        key = os.path.join(s3_prefix, csv_filename) if s3_prefix else csv_filename
        url = presigned_url(s3_bucket, key)
        if url:
            return jsonify({'url': url})
        else:
            return "File not available", 404
    try:
        logger.info('Download requested for %s', csv_filename)
        return send_from_directory(app.config['EXPORTS_FOLDER'], csv_filename, as_attachment=True)
    except FileNotFoundError:
        logger.warning('File not found for download: %s', csv_filename)
        return "File not found", 404


@app.route('/api/diagnostics', methods=['GET'])
def diagnostics():
    info = {
        'python': sys.version.splitlines()[0],
        'platform': platform.platform(),
        'tesseract': tess_path,
        'redis': bool(REDIS_URL),
        's3_bucket': os.environ.get('S3_BUCKET'),
        'memory_mb': psutil.virtual_memory().total // (1024 * 1024),
        'disk_free_mb': shutil.disk_usage(BASE_DIR).free // (1024 * 1024)
    }
    return jsonify(info)


@app.route('/nojs-validate', methods=['GET', 'POST'])
def nojs_validate():
    # Simple fallback for users without JavaScript: render a page with a basic upload form
    if request.method == 'GET':
        return '''
        <h2>Upload (no-JS fallback)</h2>
        <form method="post" enctype="multipart/form-data">
          <input type="file" name="file">
          <input type="submit" value="Upload and validate">
        </form>
        '''
    # POST handling: reuse the existing API behavior but present an HTML response
    if 'file' not in request.files:
        return '<p>No file uploaded</p>', 400
    file = request.files['file']
    if file.filename == '' or not allowed_file(file.filename):
        return '<p>Invalid file</p>', 400
    filename = secure_filename(file.filename)
    upload_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(upload_path)
    progress_key = str(uuid.uuid4())
    set_progress(progress_key, percent=0, csv_filename=None, done=False)
    # Run validation synchronously for no-JS users (short blocking)
    try:
        csv_path, _, _, _, _ = validate_pdf(upload_path, EXPORTS_FOLDER, progress_key, progress_key)
        csvfn = sanitize_csv_filename(os.path.basename(csv_path))
        return f'<p>Validation complete. <a href="/download/{csvfn}">Download CSV</a></p>'
    except Exception as e:
        return f'<p>Validation failed: {str(e)}</p>', 500

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 3000))
    app.run(host='0.0.0.0', port=port)